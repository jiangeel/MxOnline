# -*-coding=utf-8-*-
__author__ = 'Eeljiang'
__date__ = '05/02/2018 19:06'

from .models import Course, Lesson, Video, CourseResource, BannerCourse
import xadmin
from organization.models import CourseOrg


# 可以在添加课程时快速添加课程
class LessonInline(object):
    model = Lesson
    extra = 0

# 可以在添加课程时快速添加课程资源
class CourseResourseInline(object):
    model = CourseResource
    extra = 0


class CourseAdmin(object):
    # 可以传入函数 get_lesson_nums、go_to当作参数
    # list_display = ['course_name', 'course_desc', 'detail', 'degree',
    #                 'learn_times', 'students',
    #                 'fav_nums', 'image', 'click_nums', 'get_lesson_nums', 'go_to', 'add_time']
    # search_fields = ['course_name', 'course_desc', 'detail', 'degree',
    #                 'learn_times', 'students',
    #                 'fav_nums', 'image', 'click_nums']
    # list_filter = ['course_name', 'course_desc', 'detail', 'degree',
    #                 'learn_times', 'students',
    #                 'fav_nums', 'image', 'click_nums', 'add_time']
    inlines = [LessonInline, CourseResourseInline ]

    # xadmin可以识别的，指明detail的样式
    style_fields = {'detail': "ueditor"}

    # 选择加载导入excel插件
    import_excel = True

    # admin后台可以选择自动刷新时间
    refresh_times = [3, 5]

    # admin后台可以直接编辑
    list_editable = ['name', 'desc', 'detail']


    # 新增和修改Course的时候都会走这个接口
    def save_models(self):
        # 在保存课程的时候统计课程机构的课程数
        obj = self.new_obj
        obj.save()
        if obj.course_org is not None:
            course_org = obj.course_org
            course_org.course_nums = Course.objects.filter(course_org=course_org).count()
            course_org.save()

    # 操作上传后的excel文件
    def post(self, request, *args, **kwargs):
        # python3 用openpyxl
        from openpyxl import Workbook, load_workbook
        if 'excel' in request.FILES:
            # 先保存到本地
            with open('tmp.xlsx', 'wb') as f:
                for chunk in request.FILES['excel'].chunks():
                    f.write(chunk)
            # load工作簿
            wb = load_workbook('tmp.xlsx')
            # get第一张工作表名称
            sheet0 = wb.sheetnames[0]
            # 工作表对象
            worksheet = wb[sheet0]
            # 所有行对象 可迭代
            rows = worksheet.rows

            for i, row in enumerate(rows):
                # 跳过第一行标题
                if i == 0:
                    continue
                course = Course()
                course.course_org = CourseOrg.objects.get(name=row[0].value)
                course.name = row[1].value
                course.save()
        # 返回必须是以下
        return super(CourseAdmin, self).post(request, args, kwargs)


# 同一个model注册第二个管理器
class BannerCourseAdmin(object):
    list_display = ['name', 'desc', 'detail', 'degree',
                    'learn_times', 'students',
                    'fav_nums', 'image', 'click_nums', 'add_time']
    search_fields = ['name', 'desc', 'detail', 'degree',
                    'learn_times', 'students',
                    'fav_nums', 'image', 'click_nums']
    list_filter = ['name', 'desc', 'detail', 'degree',
                    'learn_times', 'students',
                    'fav_nums', 'image', 'click_nums', 'add_time']
    inlines = [LessonInline, CourseResourseInline ]


    # 重载此方法，只显示model中特定的数据
    def queryset(self):
        qs = super(BannerCourseAdmin, self).queryset()
        qs = qs.filter(degree='cj')
        return qs


class LessonAdmin(object):
    list_display = ['course', 'name', 'add_time']
    search_fields = ['course', 'name']
    list_filter = ['course__name', 'name', 'add_time']


class VideoAdmin(object):
    list_display = ['lesson', 'name', 'add_time']
    search_fields = ['lesson', 'name']
    list_filter = ['lesson__name', 'name', 'add_time']


class CourseResourceAdmin(object):
    list_display = ['Course', 'name', 'download', 'add_time']
    search_fields = ['Course', 'name', 'download']
    list_filter = ['Course__name', 'name', 'download', 'add_time']


xadmin.site.register(Course, CourseAdmin)
xadmin.site.register(BannerCourse, BannerCourseAdmin)
xadmin.site.register(Lesson, LessonAdmin)
xadmin.site.register(Video, VideoAdmin)
xadmin.site.register(CourseResource, CourseResourceAdmin)
